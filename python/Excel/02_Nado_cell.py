from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# ws["A1"] = 1
# ws["A2"] = 2
# ws["A3"] = 3

# ws["B1"] = 4
# ws["B2"] = 5
# ws["B3"] = 6

# print(ws["A1"])
# print(ws["A1"].value)
# print(ws["A10"].value)

# # row = 1, 2, 3, ...
# # column = A(1), B(2), C(3), ...
# print(ws.cell(row=1, column=1).value)
# print(ws.cell(row=1, column=2).value)

# c = ws.cell(column=3, row=1, value=10)
# print(c.value)

index = 1
# 반복문을 이용해서 랜덤 숫자 채우기
for x in range(1, 11):
    for y in range(1, 11):
        # ws.cell(row=x, column=y, value=randint(0, 100)) # 0~100 사이의 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1
wb.save("sample.xlsx")


