# import random

# Column1 = list(range(1,10))
# Column2 = list(random.randrange(4,8,2) for _ in range(len(Column1)))


# from xlsxwriter import Workbook
# workbook = Workbook('Ecl.xlsx')
# Report_Sheet=workbook.add_worksheet()
# Report_Sheet.write(0, 0, 'Column1')
# Report_Sheet.write(0, 1, 'Column2')

# for row_ind, row_value in enumerate(zip(Column1, Column2)):
#     print (row_ind, row_value)
#     for col_ind, col_value in enumerate(row_value):
#         Report_Sheet.write(row_ind + 1, col_ind, col_value)

# workbook.close()

"""
빵형 exel 파일 읽기
"""

from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성

ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" 

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2)

new_ws = wb["NewSheet"] # Dict 형태로 Sheet에 접근
print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")
